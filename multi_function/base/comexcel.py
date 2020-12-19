# -*- coding:utf-8 -*-
# author: lupengliang
from base.osusing import OSClass
from logger.logging_setting import logger
import openpyxl
from openpyxl.styles import PatternFill

# 操作excel的基层方法
class excel:
    def __init__(self):
        self.oc = OSClass()

    # 创建一个excel
    def create_excel(self, excel_name, title, data: list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        ws.append(data)
        wb.save(excel_name)
        logger.info(f'successful to create {excel_name.split("/")[-1]}.')
        return ws

    # 打开一个excel
    def open_excel(self, excel_name):
        wb = openpyxl.load_workbook(excel_name)
        ws = wb.active
        logger.info(f'successful to open {excel_name.split("/")[-1]}')
        return ws

    # 向一个excel中写入一行数据
    def add_row_data(self, ws, data: list):
        ws.append(data)
        logger.info(f'successful to write {data} to excel')

    # 向一个excel中写入一列数据
    def add_col_data(self):
        pass

    # 向一个excel的单元格写入数据
    def add_cell_data(self):
        pass

    # 将两个excel中的行数补齐
    def full_rows(self):
        pass

    # 将两个excel中的列数补齐
    def full_cols(self):
        pass

if __name__ == '__main__':
    ex = excel()
    ex.create_excel(excel_name='helloworld.xlsx', title='考试成绩表', data=['姓名', '姓别', '年龄', 'php', 'java', 'c', 'c++', 'html', 'css', 'js', 'python'])
