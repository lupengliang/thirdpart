# -*- coding:utf-8 -*-
# author: lupengliang
from base.osusing import OSClass
from logger.logging_setting import logger
import openpyxl
from openpyxl.styles import PatternFill

# class of handling excel
class excel:
    def __init__(self):
        self.oc = OSClass()

    # create one excel
    def create_excel(self, excel_name, title, data: list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        ws.append(data)
        wb.save(excel_name)
        logger.info(f'successful to create {excel_name.split("/")[-1]}.')
        return ws

    # open one excel
    def open_excel(self, excel_name):
        wb = openpyxl.load_workbook(excel_name)
        ws = wb.active
        logger.info(f'successful to open {excel_name.split("/")[-1]}')
        return ws

    # write one row data to one excel
    def add_row_data(self, ws, data: list):
        ws.append(data)
        logger.info(f'successful to write {data} to excel')

    # get max row and col min row and col from two excel
    def get_border(self, excel_name1, excel_name2):
        wb1 = openpyxl.load_workbook(excel_name1)
        ws1 = wb1.active
        wb2 = openpyxl.load_workbook(excel_name2)
        ws2 = wb2.active
        return max(ws1.max_row, ws1.max_row), min(ws1.max_row, ws2.max_row), max(ws1.max_column, ws1.max_column), min(ws1.max_column, ws2.max_column)

    # write one col data to one excel
    def add_col_data(self, excel_name):
        wb = openpyxl.load_workbook(excel_name)
        ws = wb.active
        for i in range(10):pass  # no finish, continue...
        wb.save(excel_name)


    # 向一个excel的单元格写入数据
    def add_cell_data(self):
        pass

    # 将两个excel中的行数补齐
    def full_rows(self):
        pass

    # 将两个excel中的列数补齐
    def full_cols(self):
        pass

if __name__ == '__main__':
    ex = excel()
    # ex.create_excel(excel_name='helloworld.xlsx', title='score', data=['username', 'sex', 'age', 'php', 'java', 'c', 'c++', 'html', 'css', 'js', 'python'])
    ret = ex.get_border('./helloworld1.xlsx', './helloworld2.xlsx')
    print(ret)