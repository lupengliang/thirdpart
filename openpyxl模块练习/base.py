# !/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:lupengliang
import logging

import openpyxl
from openpyxl.styles import PatternFill


# 单个excel的操作
class Excel:
    def __init__(self, excel_name, color='45b97c'):
        self.excel_name = excel_name
        self.color = color
        self.fill = PatternFill("solid", fgColor=self.color)

    # 创建一个excel
    @classmethod
    def create_excel(cls, excel_name, sheet_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(excel_name)
        logging.info(f'successful to create {excel_name.split("/")[-1]}.')
        return ws

    # 创建一个sheet页
    def create_sheet(self, wb, sheet_name, index=None):
        try:
            wb.create_sheet(sheet_name, index)
            logging.info(f'successful to create sheet named {sheet_name}.')
            return wb
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 加载一个excel
    def load_excel(self):
        try:
            wb = openpyxl.load_workbook(filename=self.excel_name)
            logging.info(f'successful to open excel named {self.excel_name}.')
            return wb
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 获取指定列头一列数据  -- 不包含列头
    def read_col_data(self, ws, col_head, max_row=None, max_col=None):
        try:
            col_index = self.read_first_row(ws=ws).index(col_head)
            for col in ws.iter_cols(min_row=2, min_col=col_index + 1, max_row=max_row, max_col=max_col):
                return [c.value for c in col]
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 获取指定行头的一行数据 -- 不包含行头
    def read_row_data(self, ws, row_head, max_row=None, max_col=None):
        try:
            row_index = self.read_first_col(ws=ws).index(row_head)
            for row in ws.iter_rows(min_row=row_index + 1, min_col=2, max_row=max_row, max_col=max_col):
                return [r.value for r in row]
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 获取行头 -- 第一列
    def read_first_col(self, ws, max_row=None, max_col=None):
        try:
            for col in ws.iter_cols(min_row=1, min_col=1, max_row=max_row, max_col=max_col):
                logging.info(f'successful to read first data on {self.excel_name}.')
                return [c.value for c in col]
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 获取列头 -- 第一行
    def read_first_row(self, ws, max_row=None, max_col=None):
        try:
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=max_col):
                logging.info(f'successful to read first data on {self.excel_name}.')
                return [r.value for r in row]
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 向单元格中通过末尾追加一行数据
    def write_by_append(self, ws, data: list):
        try:
            ws.append(data)
            logging.info(f"successful to write {data} in {self.excel_name}.")
            return ws
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 通过索引向单元格中写入数据
    def write_by_index(self, ws, x, y, data: str):
        try:
            ws.cell(x, y).fill = self.fill
            ws.cell(x, y).value = ''
            ws.cell(x, y).value = data
            logging.info(f"successful to write {data} in {self.excel_name}.")
            return ws
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 通过行头/列头进行写入单元格数据
    def write_by_head(self, ws, row_head, col_head, data):
        try:
            x = self.read_first_col(ws).index(row_head)
            y = self.read_first_row(ws).index(col_head)
            ws.cell(x, y).fill = self.fill
            ws.cell(x, y).value = ''
            ws.cell(x, y).value = data
            logging.info(f'successful to write {data} in {self.excel_name}.')
            return ws
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')

    # 保存excel
    def save_excel(self, wb, excel_name=None):
        try:
            if excel_name:
                wb.save(excel_name)
                logging.info(f"successful to save excel named {excel_name}.")
            wb.save(self.excel_name)
            logging.info(f"successful to save excel named {self.excel_name}.")
        except PermissionError:
            logging.error(f'please retry after close {self.excel_name}.')


if __name__ == '__main__':
    Excel('hello.xlsx')
