# -*- coding:utf-8 -*-
# author:lupengliang
import os

import openpyxl
from log.logger import logger


class ExcelOperation:
    def __init__(self):
        self.path = os.path.dirname(os.path.abspath(__file__))

    # 创建一个excel
    def create_excel(self, excel_name):
        wb = openpyxl.Workbook(excel_name)
        logger.info(f'successful to create excel named {excel_name}.')
        wb.active
        logger.info(f'current all sheet names are {wb.sheetnames}.')
        wb.save(excel_name)
        logger.info(f'successful to save {excel_name}.')

    # 加载已经存在的excel并写入数据
    def load_excel(self, excel_name, data: list):
        wb = openpyxl.load_workbook(excel_name)
        logger.info(f'successful to load excel named {excel_name}.')
        ws = wb.active
        ws.append(data)
        logger.info(f'successful to add data -- {data}.')

    # 删除存在的excel
    def delete_excel(self, excel_name):
        if excel_name in os.listdir(self.path):
            os.remove(excel_name)
            logger.info(f'successful to delete excel named {excel_name}.')
        else:
            logger.error(f'failed to find file named {excel_name} in {self.path}.')


if __name__ == '__main__':
    eo = ExcelOperation()